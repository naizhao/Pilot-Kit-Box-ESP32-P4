#!/usr/bin/env python3
"""把 BOM 核对清单排成能打印的 xlsx 和 pdf。

CHECKLIST.md 适合在屏幕上看和进版本控制，但打出来不好用——Markdown 表格没有
分页控制、表头不重复、勾选框也小得没法下笔。这份专门为「打印出来拿在手上，
边贴边打勾」做排版。

## 为什么要两个脚本

读 .kicad_pcb 必须用 KiCad 自带的 python（只有它有 pcbnew），而 openpyxl 装在
系统 python 里，两个解释器凑不到一起。所以 gen_checklist.py 负责从 PCB 捞数据、
落成 build/checklist.json，本脚本只管排版。

## 排版上的取舍

  · **PDF 用 A4 纵向**——横向每页只剩 25 行要 6 页且右侧留白，纵向 44 行 4 页
    就够；清单是拿在手上一行行扫的，少翻两页比行距宽松实在。xlsx 则用横向，
    因为 Excel 的列宽机制不同，横向才不会把字缩到看不清
  · **表头每页重复**——156 行要跨 4 页，翻到第 3 页还得知道哪列是什么
  · **勾选框留 6mm 见方**——圆珠笔打勾要落得下笔，太小会画到格子外
  · **按位号分组换行**——C 段结束空一行再进 R 段，扫的时候不容易串行
  · **不贴的行整行灰掉**——它们占了 20 行，灰掉后视觉上自动跳过

## 颜色

  红底  板上丝印印错了，别信板上的字（V3.2 实物板）
  黄底  有极性/方向，贴反不工作
  灰底  不贴

用法：**系统 python3** tools/gen_checklist_print.py
      先跑 gen_checklist.py（KiCad python）生成 build/checklist.json
"""
import json
import os
import subprocess
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("需要 openpyxl：pip3 install --user openpyxl")

T = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(T, "build", "checklist.json")
XLSX = os.path.join(T, "CHECKLIST.xlsx")
PDF = os.path.join(T, "CHECKLIST.pdf")

if not os.path.exists(SRC):
    sys.exit(f"没有 {SRC}\n先用 KiCad 的 python3 跑 tools/gen_checklist.py")

data = json.load(open(SRC))
rev, rows = data["rev"], data["rows"]

RED = "FFD6D6"      # 丝印印错
YELLOW = "FFF2CC"   # 有极性
GREY = "E8E8E8"     # 不贴
HEAD = "4472C4"

# ── xlsx ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = f"{rev} 核对清单"

thin = Side(style="thin", color="999999")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(["贴", "位号", "值 / 型号", "封装", "位置 (x, y)", "阶段", "备注"])
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor=HEAD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = box

prev_prefix = None
for r in rows:
    # 位号前缀换了就空一行——C 段和 R 段之间有个视觉间隔，扫的时候不容易串行
    prefix = "".join(ch for ch in r["ref"] if ch.isalpha())
    if prev_prefix and prefix != prev_prefix:
        ws.append([])
    prev_prefix = prefix

    ws.append(["", r["ref"], r["val"], r["fp"], r["pos"], r["stage"], r["note"]])
    row = ws[ws.max_row]
    fill = None
    if not r["place"]:
        fill = GREY
    elif r["silk_wrong"]:
        fill = RED
    elif r["polarity"]:
        fill = YELLOW
    for i, c in enumerate(row):
        c.border = box
        c.font = Font(size=10, bold=(i == 1))     # 位号列加粗，最常被扫
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
    row[0].alignment = Alignment(horizontal="center")
    row[1].alignment = Alignment(horizontal="center")
    row[4].alignment = Alignment(horizontal="center")
    row[6].alignment = Alignment(wrap_text=True, vertical="center")

for w, ch in zip((6, 10, 20, 26, 14, 18, 42), "ABCDEFG"):
    ws.column_dimensions[ch].width = w
ws.row_dimensions[1].height = 22
for i in range(2, ws.max_row + 1):
    ws.row_dimensions[i].height = 19          # ≈6mm，圆珠笔打勾落得下

# 打印设置：不设这些，打出来会是竖版 + 表头只在第一页 + 表格被截断
ws.print_title_rows = "1:1"                   # 每页重复表头
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0                 # 高度不限，宽度压成 1 页
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_options.gridLines = False
ws.page_margins.left = ws.page_margins.right = 0.4
ws.page_margins.top = ws.page_margins.bottom = 0.5
ws.freeze_panes = "A2"
wb.save(XLSX)
print(f"✓ {os.path.relpath(XLSX, T)}   {len(rows)} 行，A4 横向，表头每页重复")

# ── pdf（HTML → wkhtmltopdf）────────────────────────────────────────
# 走 HTML 而不是从 xlsx 转：本机没有 LibreOffice，而 wkhtmltopdf 对
# thead 每页重复、CSS 分页控制的支持比表格转换器更可靠。
def esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace("`", ""))


trs = []
prev_prefix = None
for r in rows:
    prefix = "".join(ch for ch in r["ref"] if ch.isalpha())
    if prev_prefix and prefix != prev_prefix:
        trs.append('<tr class="gap"><td colspan="7"></td></tr>')
    prev_prefix = prefix
    cls = ("skip" if not r["place"] else
           "silk" if r["silk_wrong"] else
           "pol" if r["polarity"] else "")
    trs.append(
        f'<tr class="{cls}"><td class="chk"></td><td class="ref">{esc(r["ref"])}</td>'
        f'<td class="val">{esc(r["val"])}</td><td class="fp">{esc(r["fp"])}</td>'
        f'<td class="pos">{esc(r["pos"])}</td><td class="stg">{esc(r["stage"])}</td>'
        f'<td class="note">{esc(r["note"])}</td></tr>')

n_place = sum(1 for r in rows if r["place"])
html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><style>
/* 纸张和边距在 wkhtmltopdf 的命令行参数里设——CSS 的 @page 它不认，
   写在这儿只会让人以为改了有用。 */
body {{ font-family: -apple-system, "PingFang SC", sans-serif; font-size: 9pt; }}
h1 {{ font-size: 14pt; margin: 0 0 2mm; }}
.sub {{ font-size: 8pt; color: #555; margin: 0 0 3mm; }}
.legend span {{ display: inline-block; padding: 1mm 2mm; margin-right: 3mm;
                border: 0.3mm solid #999; font-size: 8pt; }}
table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
thead {{ display: table-header-group; }}   /* 表头每页重复 */
tr {{ page-break-inside: avoid; }}
th {{ background: #4472C4; color: #fff; padding: 1.5mm 1mm; font-size: 9pt;
      border: 0.2mm solid #888; }}
td {{ border: 0.2mm solid #aaa; padding: 1mm 1mm; height: 5.5mm;
      overflow: hidden; }}
/* 列宽显式分配，加起来 196mm = A4 纵向 210 减去左右边距。
   不显式给（配合上面的 table-layout: fixed），浏览器会按内容自动分，
   备注列会被挤到折行。 */
.chk  {{ width: 9mm; }}                    /* 勾选框，6mm 见方够圆珠笔下笔 */
.ref  {{ width: 14mm; font-weight: bold; text-align: center; }}
.val  {{ width: 20mm; }}
.fp   {{ width: 40mm; font-size: 7.5pt; color: #444; }}
.pos  {{ width: 23mm; text-align: center; font-family: monospace; font-size: 7.5pt; }}
.stg  {{ width: 28mm; font-size: 8pt; }}
.note {{ width: 62mm; font-size: 7.5pt; }}
.silk {{ background: #{RED}; }}
.pol  {{ background: #{YELLOW}; }}
.skip {{ background: #{GREY}; color: #777; }}
.gap td {{ border: none; height: 2mm; padding: 0; }}
</style></head><body>
<h1>{rev} BOM 核对清单（按位号顺序）</h1>
<p class="sub">共 {len(rows)} 个位号，其中 {n_place} 个需要贴片。
贴完一个在左侧方框打勾。坐标以 PCB 左上角为原点，单位 mm。</p>
<p class="legend">
<span style="background:#{RED}">板上丝印印错了 —— 别信板上的字，认坐标</span>
<span style="background:#{YELLOW}">有极性/方向 —— 贴反不工作</span>
<span style="background:#{GREY}">不贴</span></p>
<table><thead><tr><th>贴</th><th>位号</th><th>值 / 型号</th><th>封装</th>
<th>位置 (x, y)</th><th>阶段</th><th>备注</th></tr></thead>
<tbody>{''.join(trs)}</tbody></table></body></html>"""

tmp_html = os.path.join(T, "build", "checklist.html")
open(tmp_html, "w").write(html)
# 纸张方向走命令行参数——CSS 的 `@page { size: ... }` wkhtmltopdf 不认。
#
# 用**纵向**：实测横向每页只剩 25 行、要 6 页，右侧还大片留白；纵向每页 44 行、
# 4 页就够。核对清单是拿在手上一行行扫的，少翻两页比行距宽松更实在。
# 列宽已在 CSS 里显式分配，纵向也放得下长备注不折行。
r = subprocess.run(["wkhtmltopdf", "--quiet", "--enable-local-file-access",
                    "--page-size", "A4", "--orientation", "Portrait",
                    "--margin-top", "8mm", "--margin-bottom", "8mm",
                    "--margin-left", "7mm", "--margin-right", "7mm",
                    tmp_html, PDF], capture_output=True, text=True)
if r.returncode == 0:
    print(f"✓ {os.path.relpath(PDF, T)}   A4 纵向（打印用）")
else:
    print(f"✗ PDF 生成失败（xlsx 已生成，可从 Excel 打印）:\n{r.stderr[:300]}")
