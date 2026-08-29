#!/usr/bin/env python3
"""生成嘉立创 SMT 贴片用的 BOM（按官方模板的四列格式）。

跟 gen_bom_xlsx.py（发给淘宝卖家的购物清单）**是两份完全不同的东西**：

                    购物清单                    SMT BOM
    读者            淘宝卖家（人）                嘉立创的匹配系统（机器）
    位号            不需要，已删                  **必须完整**，贴片机按它下料
    已买的料         剔除                        **全部列上**，否则那些位置没人贴
    合并            按规格合并成一种料买            按 (值,封装) 分组，位号全列
    品牌            参考，可换                    要给立创料号，不然匹配不上

## 嘉立创模板的四列（来自官方 ___BOM______.xls）

    Comment型号 | Designator位号 | Footprint封装 | 嘉立创元件编号(可不要此列)

模板里三条硬规则，踩了就翻车：
  · 位号之间**只能用逗号或空格**分隔，不能用顿号/冒号/反斜杠/竖线
  · 区间表达只支持 `C3-5` 这种；本脚本一律全展开，不用区间，省得歧义
  · **位号列绝对不能写备注**——模板原话「写备注一定会翻车」

## 数据来源

位号和分组直接复用 gen_bom.py（它从网表读，有完整位号；只是打 Markdown 时
截断到 6 个显示，数据本身是全的）。立创料号两个来源：
  · gen_bom.py 的 PARTS 表（选定料的那些）
  · 下面的 ORDERED_LCSC（主芯片，来自集成扩展板设计（内部文档）的三渠道总表）

用法：**系统 python3** tools/gen_bom_smt.py

⚠️ 别用 KiCad 自带的解释器。本仓库两个 python 混用：
    KiCad 的 (3.9)   —— 只有它有 pcbnew，凡是要读写 .kicad_pcb 的脚本必须用它
    系统的  (3.13)   —— 有 openpyxl 等第三方库
本脚本只读网表 XML、写 xlsx，属于后者。拿 KiCad 的解释器跑会卡在缺 openpyxl，
而且 3.9 的 f-string 表达式里不允许反斜杠，一不小心就 SyntaxError。
"""
import os
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("需要 openpyxl：pip3 install --user openpyxl")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
T = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# gen_bom.py 读 /tmp/expansion.net.xml——/tmp 会被系统清，隔几天再跑就 FileNotFound。
# 不在就自己导一份，别让人去猜该跑哪条 kicad-cli。
_NET = "/tmp/expansion.net.xml"
if not os.path.exists(_NET):
    import subprocess
    _CLI = os.path.expanduser("~/Applications/KiCad/KiCad.app/Contents/MacOS/kicad-cli")
    subprocess.run([_CLI, "sch", "export", "netlist", "--format", "kicadxml", "-o", _NET,
                    os.path.join(T, "kicad", "expansion-board-v3.kicad_sch")],
                   check=True, capture_output=True)
    print(f"（网表不在，已重新导出 {_NET}）")

import gen_bom as G                                   # noqa: E402  会顺带重写 internal/BOM_PURCHASE.md（幂等）

# v3 已归档：贴片下单 BOM 属内部文档 → 落 internal/
OUT = os.path.join(T, "internal", "BOM_嘉立创SMT.xlsx")

# ── 主芯片的立创料号 ────────────────────────────────────────────────
# gen_bom.py 里这些标成 HAVE（2026-08-01 已自购），料号栏是「—」。
# 但 SMT 报价要看**能不能配上料**，没料号嘉立创就匹配不上，等于白问。
# 料号全部抄自集成扩展板设计（内部文档）的
# 三渠道总表，不是猜的。
ORDERED_LCSC = {
    "CC1312R1F3RGZR":     "C543014",     # §7 行128，现货仅 35
    "RP2040":             "C2040",       # §7 行130
    "W25Q128JVSIQ":       "C97521",      # §7 行131
    "QPL9547TR7":         "C5367093",    # §7 行135
    "BGA2817":            "C2654187",    # §7 行137
    "TA0970A":            "C7115531",    # §7 行138
    "AD8319ACPZ-R7(主)":   "C21652",      # §7 行139
    "AD8313ARMZ(备,二选一)": "C578690",     # §7 行140，现货仅 5
    "TLV3501(TOKMAS)":    "C54582190",   # §7 行141，TI 原装是假货重灾区
    "BNO085":             "C5189642",    # §7 行142，订货
    "BMP388":             "C779278",     # §7 行144
    "QMC5883P":           "C2847467",    # §7 行145
    "ATGM336H-6N-74":     "C5804601",    # §7 行146，现货仅 8
    "ME6211C33M5":        "C82942",      # §7 行148（料号是 ME6211C33M5G-N）
    "TPS7A2033PDBVR":     "C2862740",    # §7 行149
}

# ── 通用阻容的嘉立创料号 ────────────────────────────────────────────
# 全部实查自嘉立创元件库接口（2026-08-15），不是猜的：
#   POST jlcpcb.com/api/overseas-pcb-order/v1/shoppingCart/smtGood/selectSmtComponentList
#
# **优先选基础库（base）**：嘉立创对扩展库元件每种收一次上料费，样板只做 5 片时
# 这笔钱比料本身贵得多。下面 23 项里 20 项走基础库，只有 3 项基础库确实没有。
#
# ⚠️ 耐压跟发给淘宝的购物清单不完全一致——**基础库的耐压反而更高**，比如
# 1uF/0402 淘宝建议 16V、基础库直接是 25V；22uF/0805 建议 16V、基础库是 25V
# 且库存 394 万（扩展库同规格只有 1.4 万）。耐压更高只会更好（DC bias 降容更小），
# 不是冲突，别去"对齐"成低的那个。
#
#   (值, 嘉立创封装) → (料号, 实际规格, 是否基础库)
BASIC_LCSC = {
    ("100nF", "0402"): ("C307331", "±10% 50V X7R",    True),   # Samsung CL05B104KB54PNC
    ("100pF", "0402"): ("C1546",   "±5% 50V C0G",     True),   # 风华 0402CG101J500NT
    ("1uF",   "0402"): ("C52923",  "±10% 25V X5R",    True),   # Samsung CL05A105KA5NQNC
    ("1k",    "0402"): ("C11702",  "±1% 62.5mW",      True),   # UNI-ROYAL 0402WGF1001TCE
    ("100nF", "0603"): ("C14663",  "±10% 50V X7R",    True),   # YAGEO CC0603KRX7R9BB104
    ("100pF", "0603"): ("C14858",  "±5% 50V C0G",     True),   # Samsung CL10C101JB8NNNC
    ("1uF",   "0603"): ("C15849",  "±10% 50V X5R",    True),   # Samsung CL10A105KB8NNNC
    ("12pF",  "0603"): ("C38523",  "±5% 50V C0G",     True),   # Samsung CL10C120JB8NNNC
    ("15pF",  "0603"): ("C1644",   "±5% 50V C0G",     True),   # Samsung CL10C150JB8NNNC
    ("1nF",   "0603"): ("C1588",   "±10% 50V X7R",    True),   # Samsung CL10B102KB8NNNC
    ("200pF", "0603"): ("C113796", "±5% 50V NP0",     False),  # YAGEO CC0603JRNPO9BN201，基础库无
    ("3pF",   "0603"): ("C46219",  "±0.25pF 50V C0G", False),  # 风华 0603CG3R0C500NT，基础库无
    ("10uF",  "0805"): ("C15850",  "±10% 25V X5R",    True),   # Samsung CL21A106KAYNNNE
    ("22uF",  "0805"): ("C45783",  "±20% 25V X5R",    True),   # Samsung CL21A226MAQNNNE
    ("0R",    "0603"): ("C21189",  "±1% 100mW",       True),   # UNI-ROYAL 0603WAF0000T5E
    ("100k",  "0603"): ("C25803",  "±1% 100mW",       True),
    ("10k",   "0603"): ("C25804",  "±1% 100mW",       True),
    ("18k",   "0603"): ("C25810",  "±1% 100mW",       True),
    ("1k",    "0603"): ("C21190",  "±1% 100mW",       True),
    ("27R",   "0603"): ("C25190",  "±1% 100mW",       False),  # 基础库无 27Ω 这个阻值
    ("4.7k",  "0603"): ("C23162",  "±1% 100mW",       True),
    ("5.1k",  "0603"): ("C23186",  "±1% 100mW",       True),
    ("USB-C_16P", "TYPE-C-31-M-12"): ("C165948", "Type-C 母 卧贴 16P", False),  # HRO 原厂，跟封装同名
}

# ── 不贴片的位置 ──────────────────────────────────────────────────
# 列进 SMT BOM 会让嘉立创去配一份根本不焊的料，或者直接报「匹配不到」。
SKIP_FP = ("TestPoint_Pad", "SolderJumper", "ANT_IFA")   # 焊盘/天线，本来就无器件
SKIP_VAL = re.compile(r"DNP")                            # 设计上默认不贴
# AD8313 是「A/B 实验二选一」的备用检波位，默认贴 AD8319(U13)。两颗都贴会打架，
# 而且 AD8313 单价 71 元，白白抬高报价。
SKIP_REF = {"U14"}

# ── KiCad 封装名 → 嘉立创封装名 ─────────────────────────────────────
# 有立创料号时这列只是辅助（嘉立创按料号走），没料号的通用阻容全靠它匹配。
FP_MAP = [
    (r"^[CRLD]_(\d{4})_\d+Metric$",       r"\1"),          # C_0402_1005Metric → 0402
    (r"^Fuse_(\d{4})_\d+Metric$",         r"\1"),
    (r"^D_SOD-(\d+)$",                    r"SOD-\1"),
    (r"^SOT-363_SC-70-6$",                "SOT-363"),
    (r"^SOIC-8_5\.3x5\.3mm_P1\.27mm$",    "SOIC-8_208mil"),   # 5.3mm 体宽 = 208mil 宽体
    (r"^MSOP-8_3x3mm_P0\.65mm$",          "MSOP-8_3x3mm_P0.65mm"),
    (r"^QFN-48-1EP_7x7mm_P0\.5mm.*",      "QFN-48_7x7mm_P0.5mm"),
    (r"^QFN-56-1EP_7x7mm_P0\.4mm.*",      "QFN-56_7x7mm_P0.4mm"),
    (r"^DFN-8-1EP_2x2mm_P0\.5mm.*",       "DFN-8_2x2mm_P0.5mm"),
    (r"^Crystal_SMD_3225-4Pin.*",         "SMD3225-4P"),
    (r"^Crystal_SMD_3215-2Pin.*",         "SMD3215-2P"),
    (r"^U\.FL_Hirose.*",                  "IPEX-1"),          # 立创对 U.FL 座的通称
    (r"^USB_C_Receptacle_HRO_(.*)$",      r"\1"),             # → TYPE-C-31-M-12
    (r"^PinHeader_2x20_P2\.54mm.*",       "HDR-2x20_P2.54mm_SMD"),
    # 自制封装：文件名是内部命名，换成业界通用写法，尺寸取自各自的 datasheet
    (r"^BMP388_LGA-10$",                  "LGA-10_2x2mm_P0.5mm"),
    (r"^BNO085_LGA-28$",                  "LGA-28_3.8x5.2mm_P0.5mm"),
    (r"^ATGM336H_LCC-18$",                "LCC-18_10.1x9.7mm"),
    (r"^AD8319_LFCSP-8-EP$",              "LFCSP-8_3x2mm_P0.5mm"),
    (r"^TA0970A_SMD3838-6$",              "SMD3838-6P"),
]


def jlc_fp(fp):
    for pat, rep in FP_MAP:
        if re.match(pat, fp):
            return re.sub(pat, rep, fp)
    return fp                                    # LGA-16_3x3mm_P0.5mm 这类本来就通用


def comment(val, fp, mat, volt):
    """Comment 列 = 卖家/系统看的型号。

    通用阻容只写值（`100nF`）匹配不到唯一料——立创同容值有几十个 SKU。
    补上封装/耐压/材质，匹配系统才能收敛。
    ⚠️ 位号列不能写备注，但 Comment 列可以，别写反了。
    """
    v = re.sub(r"（[^）]*）|\([^)]*\)", "", val).strip()
    v = re.sub(r"\s*(串|并|旁路外接|-天线侧|1090实调)\s*", "", v).strip() or val
    if fp.startswith("C_"):
        return " ".join(x for x in (v, volt, mat) if x)
    return v


# 材质/耐压复用购物清单那套规则，两份表对同一颗电容的说法必须一致
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gen_bom_xlsx import cap_voltage, material      # noqa: E402

rows, skipped = [], []
for cat, lvl, val, fp, n, refs, part, lcsc, note in G.rows:
    if any(s in fp for s in SKIP_FP):
        skipped.append((val, refs, "PCB 焊盘/天线，无器件"))
        continue
    if SKIP_VAL.search(val):
        skipped.append((val, refs, "设计上默认不贴（DNP）"))
        continue
    refs = [r for r in refs if r not in SKIP_REF]
    if not refs:
        skipped.append((val, sorted(SKIP_REF), "A/B 实验备用位，默认贴 AD8319"))
        continue
    code = lcsc if lcsc and lcsc != "—" else ORDERED_LCSC.get(val, "")
    mat = material(part, val, note, fp)
    mat = "" if mat == "不限" else mat            # 「不限」是给人看的，给机器就成了噪音
    volt = cap_voltage(val, jlc_fp(fp)) if fp.startswith("C_") else ""
    cm = comment(val, fp, mat, volt)
    basic = None
    if not code:
        # 通用阻容：查到了嘉立创的现货料就用它，Comment 也换成**那颗料的真实规格**。
        # 不换的话表里会写着 16V、配的料是 25V，人对不上账。
        # ⚠️ 这个"取第一个词"要先算成变量，别塞进 f-string——KiCad 自带的
        # Python 是 3.9，**f-string 表达式里不允许出现反斜杠**（3.12 才放开）。
        # 写成 f"{re.sub(r'\s.*', '', cm)}" 在系统 python 3.13 上跑得好好的，
        # 换 KiCad 的解释器就直接 SyntaxError。本仓库两个解释器混用，
        # 凡是不依赖 pcbnew 的脚本都要保证 3.9 也能跑。
        head = re.sub(r"\s.*", "", cm)
        hit = BASIC_LCSC.get((head, jlc_fp(fp))) or BASIC_LCSC.get((val, jlc_fp(fp)))
        if hit:
            code, spec, basic = hit
            cm = f"{head} {spec}" if fp[0] in "CRL" else cm
    rows.append([cm, refs, jlc_fp(fp), code, basic])

# 合并同料。ZS1 在原理图上画在电感位（L_0603）、值写「0R 串」，R22/R23 画在电阻位，
# 但**买的时候就是同一颗 0603 0R**。不合并就是两行同型号同封装，嘉立创会当成两种料
# 各收一次上料费。
mg = {}
for cm, refs, fph, code, basic in rows:
    mg.setdefault((cm, fph, code, basic), []).extend(refs)


def ref_key(r):
    m = re.search(r"\d+$", r)
    return (re.sub(r"\d+$", "", r), int(m.group()) if m else 0)


rows = [[cm, ",".join(sorted(rs, key=ref_key)), fph, code, len(rs), basic]
        for (cm, fph, code, basic), rs in mg.items()]
rows.sort(key=lambda r: (ref_key(r[1].split(",")[0])[0], r[0]))

wb = Workbook()
ws = wb.active
ws.title = "BOM"
ws.append(["Comment", "Designator", "Footprint", "JLCPCB Part #（LCSC）"])
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
for r in rows:
    ws.append(r[:4])
for w, ch in zip((26, 62, 26, 16), "ABCD"):
    ws.column_dimensions[ch].width = w
for row in ws.iter_rows(min_row=2):
    row[1].alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "A2"
wb.save(OUT)

# 机检：位号总数必须对得上，不能悄悄漏件
placed = sum(r[4] for r in rows)
dropped = sum(len(s[1]) for s in skipped)
assert placed + dropped == sum(r[4] for r in G.rows), \
    f"位号对不上：贴片 {placed} + 不贴 {dropped} ≠ 网表 {sum(r[4] for r in G.rows)}"

print(f"SMT BOM：{len(rows)} 个料号 / {placed} 个贴片位 → {OUT}")
print(f"\n不贴片的 {dropped} 个位置（已排除）：")
for val, refs, why in skipped:
    print(f"    {val[:22]:22s} {','.join(refs)[:26]:26s} {why}")
nb = sum(1 for r in rows if r[5] is True)
ne = sum(1 for r in rows if r[5] is False)
print(f"\n通用阻容选料：基础库 {nb} 项（免上料费）/ 扩展库 {ne} 项")
for r in rows:
    if r[5] is False:
        print(f"    ⚠️ 扩展库 {r[0][:30]:30s} {r[2][:16]:16s} {r[3]}")
nolcsc = [r for r in rows if not r[3]]
if nolcsc:
    print(f"\n⚠️ 没有立创料号的 {len(nolcsc)} 项，嘉立创匹配不上，要么自供要么人工选料：")
    for r in nolcsc:
        print(f"    {r[0][:30]:30s} {r[2][:24]:24s} ×{r[4]}")
