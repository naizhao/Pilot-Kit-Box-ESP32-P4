#!/usr/bin/env python3
"""生成发给淘宝卖家的购物清单（Excel）。

2026-08-13 评审否掉了第一版："你出的这个文件更加像是给我内部看的，而不是给淘宝
卖家的。你里面备注太详细了，用在什么地方、必买之类的，根本不需要给卖家。
这不是一份购物清单。"

## 所以这份表只有卖家报价需要的东西

    型号规格 | 封装 | 数量 | 参考料号 | 规格要求

**没有**：用途、位号、采购等级、"必买/有推荐"、板上用量 vs 建议采购的对照、
避坑说明的来龙去脉。卖家不关心这些，看到一堆内部术语只会看不懂或者报错价。

「规格要求」列保留但压到最短：只写**选错就不能用**的硬指标（绕线/C0G/±1%/
低结电容），因为这类料淘宝上同名不同质的太多，不写清楚会拿到废件。

## 已买的不进表

  · 集成扩展板设计（内部文档）「BOM 三渠道
    总表」里的 20 项，2026-08-01 已按那份表下过单
  · IPEX(U.FL) 座、USB-C 座 —— 已有现货
  · 天线/测试点这类 PCB 焊盘，本来就不是器件

阻容感全部重新采购（评审要求：阻容感价格不高，全部重新采购），
所以不再区分"元件本自备"。

用法：gen_bom_xlsx.py     读 ../BOM_PURCHASE.md，写 ../BOM_采购清单.xlsx
"""
import os
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("需要 openpyxl：pip3 install --user openpyxl")

T = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# v3 已归档：采购文档都在 internal/
SRC = os.path.join(T, "internal", "BOM_PURCHASE.md")
OUT = os.path.join(T, "internal", "BOM_采购清单.xlsx")

# ── 已买的，不进清单 ──────────────────────────────────────────────────
# 来源：集成扩展板设计（内部文档）「BOM 三渠道总表」
BOUGHT_KEYWORDS = [
    "CC1312R", "RP2040", "W25Q128", "X322512", "ABM8W", "Q13FC135", "FC-135",
    "QPL9547", "BGA2817", "TA0970A", "AD8319", "AD8313", "TLV3501",
    "BNO085", "BMP388", "QMC5883", "ATGM336H", "ME6211", "TPS7A2033",
    "MM8930",                      # RF 测试座
]
# 已有现货、不用再买的
HAVE_KEYWORDS = ["U.FL", "USB_C", "TYPE-C", "排针", "PinHeader"]
# 不是器件
# 不是器件 / 不用买。DNP = Do Not Populate，板上留了位置但不贴件，
# 列进采购清单等于让卖家报一份根本不会焊上去的料。
NOT_PART = ["PCB 铜箔", "PCB 焊盘", "无器件", "DNP"]

# 规格要求：只留"选错就不能用"的，一句话说完
# 容差档（EIA）：射频匹配用的 pF 级电容，**容差比品牌重要得多**。
# 品牌只要材质和容差对得上就能换，容差错了匹配点就偏。
TOL = {"B": "±0.10pF", "C": "±0.25pF", "D": "±0.5pF", "F": "±1%", "G": "±2%",
       "J": "±5%", "K": "±10%", "M": "±20%"}

SPEC_SHORT = [
    (r"绕线", "必须绕线电感，叠层款不行"),
    # ⚠️ 晶振负载电容原文是「**建议** C0G；本子若是 X7R，频偏可接受但请实测」——
    # 是建议不是要求。写成"必须"会让卖家按更贵的料报价，也不符合设计意图。
    # ⚠️ 判据要用「建议 C0G」本身，不能用「晶振负载」——15pF 那行 note 写的是
    # 「同上：建议 C0G，本子货请实测频偏」，压根没有"晶振负载"四个字，会漏掉，
    # 结果同一件事的两颗电容一颗写"建议"一颗写"必须"。
    (r"建议\s*C0G", "建议 C0G（X7R 可用，需实测频偏）"),
    (r"C0G|NP0", "必须 C0G/NP0，不要 X7R"),
    (r"±1%", "精度 ±1%"),
    (r"结电容|Cj", "低结电容 ESD 管（≤0.6pF）"),
    (r"DCR|Isat", "DCR<300mΩ 且 Isat≥100mA"),
]


# 电容耐压。2026-08-14 下单时问到「电压怎么选」——原始 BOM 只写了容值和
# 材质，没写耐压，卖家页面上同一容值有 6.3/10/16/25/35/50V 六个档。
#
# 选档不是怕击穿（6.3V 顶 3.3V 也不会炸），是 **DC bias 降容**：X5R/X7R 在直流偏压
# 下容值会塌，0402 1µF 6.3V 加 3.3V 偏压后实测只剩 0.4~0.5µF——按 1µF 设计的去耦
# 板上只有一半。所以规则是 **耐压 ≥ 2× 工作电压**。
#
# 各档依据（网络归属实测自 kicad/expansion-board-v3.kicad_pcb）：
#   VCC_5V            C1(0805 10µF) C2 C3(0603)     → 板上最高 5V
#   3V3_DIG/RF/GNSS   绝大多数去耦                    → 3.3V
#   SUBG_VDDR ≈1.7V   C60(0805 22µF) C61 C62
#   RP_1V1 / SUBG_DCOUPL  C28 C29 / C67             → 1.1~1.28V
#   射频耦合匹配        C30~C45 C53~C59              → 交流小信号
# 同容值跨多个电压域时取**最高的那个**，合并成一种料买，省得分两包。
# C0G/NP0 没有 DC bias 效应，50V 是这类料号的标准档（料号尾部 `500` 即 50V）。
CAP_V = [
    # (封装, 容值正则, 耐压)
    ("0805", r"^22[uµ]F",  "16V"),   # 22µF/0805 常规档就到 16V，25V 极少
    ("0805", r"^10[uµ]F",  "25V"),   # C1 在 5V 轨，留裕量
    ("0603", r"^4\.7[uµ]F", "16V"),
    ("0603", r"^1[uµ]F",   "25V"),   # C2/C3 在 5V 轨，与 3V3 的合并
    ("0402", r"^1[uµ]F",   "16V"),   # 3V3_RF；6.3V 会降容一半
    ("",     r"^\d+nF",    "50V"),   # nF 级本来就小，降容无所谓，50V 最通用
    ("",     r"pF",        "50V"),   # C0G，无 bias 效应
]


def cap_voltage(val, pkg_h):
    for p, pat, v in CAP_V:
        if (not p or p == pkg_h) and re.search(pat, val, re.I):
            return v
    return ""


def material(model, val, note, pkg):
    """材质/工艺。电容看介质，电感看绕制工艺——都是**换了就不能用**的硬指标。"""
    s = f"{model} {val} {note}"
    m = re.search(r"(X5R|X7R|C0G|NP0)", s, re.I)
    if m:
        return "C0G/NP0" if m.group(1).upper() in ("C0G", "NP0") else m.group(1).upper()
    if pkg.startswith("C_"):
        # 厂商编码里的介质位，品名里没写明的靠这个兜底：
        #   三环 TCC0402**COG**3R0…  风华 0402**CG**2R7…  YAGEO …**NPO**9BN3R6
        #   Samsung CL10**A**475…（A=X5R，B=X7R）
        if re.search(r"COG|NPO|\d{4}CG\d", model, re.I):
            return "C0G/NP0"
        if re.search(r"\bCL\d+A\d", model):
            return "X5R"
    if pkg.startswith("L_"):
        # 只有原始 BOM 明确标了「绕线」的才是。板上两类电感选型逻辑完全不同：
        #   射频用（0402 那 9 颗）—— 看 Q 值和 SRF，必须绕线。叠层款 SRF 只有
        #                          600-700MHz，到 1090MHz 已经变成电容
        #   DC-DC 储能（L7 6.8uH）—— 看 DCR 和 Isat，工作频率才几 MHz，
        #                          叠层/模压都行，型号 TDK MLZ 本身就是叠层
        # 留空会让人分不清是"没要求"还是"漏填"，所以明写「不限」。
        return "绕线" if re.search(r"绕线", s) else "不限"
    # 原始 BOM 没写材质的电容（0603 的 100pF/200pF/1nF/3pF 这些模拟滤波件），
    # 留空会让卖家来问「要什么材质」。明写「不限」——原文确实没提要求，
    # 这是**如实传达**，不是替设计做主。
    return "不限" if pkg.startswith("C_") else ""


def cap_tol(model):
    """从厂商料号解出容差档：TCC0402COG3R0**C**500AT → C → ±0.25pF"""
    # 两种料号格式都要认：
    #   三环/风华  COG + 容值(3R0) + 容差(C)      TCC0402COG3R0C500AT
    #   YAGEO     NPO + 容差(9B) + N + 容值(3R6)  AC0402CRNPO9BN3R6
    m = re.search(r"(?:COG|CG|NPO)\s*(\d[R\d]\d)([BCDFGJKM])", model, re.I)
    if m:
        return TOL.get(m.group(2).upper())
    m = re.search(r"NPO\d([BCDFGJKM])N", model, re.I)
    return TOL.get(m.group(1).upper()) if m else None


def short_spec(model, note):
    s = model + " " + note
    # ⚠️ 原文里出现 "C0G" 不等于"必须 C0G"。100pF 0402 那条写的是
    # 「✅实算：X7R vs C0G 全路径只差 0.13dB，不值得单买 NP0」——是**否定**的语境。
    # 只按关键词匹配会把它标成"必须 C0G"，跟品名里的 X7R 自相矛盾，卖家看了会来问。
    # ⚠️ 否定词不能放 `可用`：52.3R 的 note 是「必须 ±1%，**不可用** 51R±5% 代」，
    # 里面含「可用」两个字，会把这条**最该写出来**的硬要求整个吞掉——真需要 ±1%
    # 的那颗规格栏空白，反倒是随便买的 1k 标着「精度 ±1%」。
    if re.search(r"不值得|X7R\s*即可", note):
        return ""
    for pat, txt in SPEC_SHORT:
        if re.search(pat, s):
            if "C0G" in txt:
                t = cap_tol(model)
                if t:
                    return f"C0G/NP0，容差 {t}"
            return txt
    return ""


# 内部说法，卖家看不懂：「0603 电容本 100nF」「任意 1k ±1%」「调试用，先贴 0R 直通」
INTERNAL = re.compile(r"电容本|电阻本|电感本|元件本|任意|调试用|DNP|待定|实测确定|上板")


def pkg_size(pkg):
    """C_0402_1005Metric → 0402"""
    m = re.search(r"_(\d{4})_", pkg)
    return m.group(1) if m else ""


# 元件类型。2026-08-14 评审要求："不要在电容里面穿插电阻，人类不好管理。"
# 按封装前缀判类型，权重决定表里的分组顺序——同类连成一片，采购时一段一段核对。
KIND = [
    (r"^C_",        "电容", 1),
    (r"^R_",        "电阻", 2),
    (r"^L_",        "电感", 3),
    (r"^D_|SOD-",   "二极管", 4),
    (r"^Fuse",      "保险丝", 5),
    (r"SOT-23$|SOT-23-\d", "三极管/MOS", 6),
    (r"SOT-363|SC-70", "IC", 7),
]


def kind_of(pkg, name=""):
    # ⚠️ 0Ω 跳线在 KiCad 里可能画成电感位（L_0603），但**买的时候就是 0603 电阻**。
    # 按封装归类会让采购在电感区里找到一个 0R，对着卖家说不清。
    if re.match(r"^0R\b", name):
        return "电阻", 2
    for pat, kname, w in KIND:
        if re.search(pat, pkg):
            return kname, w
    return "其他", 9


def pkg_human(pkg):
    """KiCad 封装名 → 卖家看得懂的。
    「C_0402_1005Metric」「Fuse_0805_2012Metric」这种是 KiCad 内部命名，
    发给卖家对方得先猜一遍。"""
    sz = pkg_size(pkg)
    if sz:
        return sz
    for pat, out in (("SOT-23", "SOT-23"), ("SOT-363", "SOT-363/SC-70-6"),
                     ("SOD-123", "SOD-123"), ("SOD-323", "SOD-323")):
        if pat in pkg:
            return out
    return pkg


def clean_model(model, val, pkg="", note=""):
    """卖家视角的标准品名——**只有规格，不带品牌**。

    2026-08-14 评审提出：「是否也按照标准的电容标识来写，但增加一列建议/参考的品牌？
    因为我的理解，不是一定要锁定这些品牌，只要能达到标准就可以了吧？」——对。
    射频匹配电容的硬指标是「材质 + 容值 + 耐压 + 容差」，牌子谁家的都行。
    品名里写死「三环 TCC0402COG3R0C500AT」，卖家会当成必须原厂，配不到就来问，
    或者干脆报个贵的。品牌挪去单独的「参考品牌/料号」列，是**建议**不是**要求**。

    材质和耐压也各自成列（material() / cap_voltage()），这里只留值本身。
    """
    v = re.sub(r"（[^）]*）|\([^)]*\)", "", val).strip()
    v2 = re.sub(r"\s*(DNP|串|并|旁路外接|-天线侧|1090实调).*", "", v).strip()
    # ⚠️ 去掉备注后可能什么都不剩（"DNP 并-天线侧" → 空），那就退回原值——
    # 生成一个空品名，卖家根本没法报价。
    return v2 if re.search(r"[\dA-Za-z]", v2) else v


def brand_ref(model, name):
    """参考品牌/料号。留空 = 这项没有品牌偏好，随便买。"""
    if not model or INTERNAL.search(model):
        return ""                       # 「任意 1k ±1%」「0603 电容本 100nF」不是品牌
    m = re.sub(r"（[^）]*）|\([^)]*\)", "", model).strip()
    return "" if m == name else m       # 跟品名一样就别在两列里重复一遍


rows = []
for line in open(SRC):
    if not line.startswith("| ") or "---" in line:
        continue
    f = [x.strip() for x in line.strip().strip("|").split("|")]
    if len(f) < 9 or f[0] == "采购":
        continue
    rows.append(f)
assert rows, f"没从 {SRC} 解析出任何行——格式变了？"

merged, skipped = {}, []
for f in rows:
    grade, cat, val, pkg, qty, refs, model, lcsc, note = f[:9]
    # ⚠️ 判「这项是不是已经买过的主芯片」只能看**型号和值**，绝不能带上「类别」。
    #
    # 类别写的是**用途**，里面经常出现芯片名。L7 的类别是 "CC1312R DCDC"
    # ——它是给 CC1312R 供电的那颗 6.8uH 电感，结果撞上 BOUGHT_KEYWORDS 里的
    # "CC1312R"，被当成「CC1312R 芯片本身，已下单」整行剔除，采购清单里就少了
    # 一颗电感。评审清点实物 BOM 时发现的。
    #
    # 主芯片行不受影响：它们的 val 和 model 本身就是型号（CC1312R1F3RGZR），
    # 不靠类别也能匹配上。
    blob = (model + " " + val).upper()
    if any(k.upper() in blob for k in NOT_PART):
        continue                                   # 不是器件，静默跳过
    if any(k.upper() in blob for k in BOUGHT_KEYWORDS):
        skipped.append((clean_model(model, val, pkg, note), qty, "§7 三渠道总表，2026-08-01 已下单"))
        continue
    if any(k.upper() in blob for k in HAVE_KEYWORDS):
        skipped.append((clean_model(model, val, pkg, note), qty, "已有"))
        continue
    name = clean_model(model, val, pkg, note)
    key = (name, pkg_human(pkg))
    if key in merged:
        merged[key][2] += int(qty)
    else:
        merged[key] = [name, pkg, int(qty), lcsc if lcsc != "—" else "",
                       short_spec(model, note), val,
                       material(model, val, note, pkg),
                       brand_ref(model, name)] + list(kind_of(pkg, name))

# 采购数量：阻容感这类几分钱的按整包，其余按用量留余量。
# 评审要求："有一些元件便宜的可以适当多采点"
CHEAP = re.compile(r"^[CRL]_\d{4}|电阻|电容|电感")


def order_qty(qty, pkg, name):
    if CHEAP.search(pkg) or CHEAP.search(name):
        return 100 if "0402" in pkg else 50
    return max(qty + 2, round(qty * 1.3))


buy = []
for (name, pkgh), m in sorted(merged.items(),
                              key=lambda z: (merged[z[0]][9], z[0][1], z[0][0])):
    nm, pk, qty, lcsc, spec, val, mat, brand, kind, _w = m
    volt = cap_voltage(val, pkgh) if kind == "电容" else ""
    buy.append([kind, nm, pkgh, mat, volt, order_qty(qty, pk, nm),
                brand, lcsc, spec])

wb = Workbook()
ws = wb.active
ws.title = "采购清单"
ws.append(["类型", "型号 / 规格", "封装", "材质", "耐压", "数量",
           "参考品牌/料号（非强制）", "立创料号（参考）", "规格要求"])
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
for r in buy:
    ws.append(r)
for w, ch in zip((11, 16, 8, 10, 7, 7, 28, 14, 26), "ABCDEFGHI"):
    ws.column_dimensions[ch].width = w
for row in ws.iter_rows(min_row=2):
    row[8].alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"采购清单 {len(buy)} 项 → {OUT}")
print(f"（已买/已有，未列入：{len(skipped)} 项）")
for n, q, why in sorted(set(skipped)):
    print(f"    {n[:34]:34s} ×{q:3s}  {why}")
